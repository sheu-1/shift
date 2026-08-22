import io
import os
import random
from datetime import datetime, date, timedelta
from types import SimpleNamespace

from flask import (
    Flask, render_template, request, redirect, url_for, flash, send_file, jsonify
)
from supabase import create_client, Client
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__)

app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "change-this-secret-key-in-production")

SUPABASE_URL = os.environ.get("SUPABASE_URL")
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_KEY")

supabase: Client = None
if SUPABASE_URL and SUPABASE_KEY:
    supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
else:
    print("WARNING: SUPABASE_URL and SUPABASE_SERVICE_KEY must be set!")

# ------------------------------------------------------------- config --
def get_roster_date_range():
    if not supabase: return None, None
    try:
        res = supabase.table("worker").select("name").ilike("name", "___CONFIG___|%").execute()
        if res.data:
            name_val = res.data[0].get("name", "")
            parts = name_val.split("|")
            if len(parts) == 3:
                return parts[1], parts[2]
    except Exception as e:
        print(f"Error fetching config: {e}")
    return None, None

def set_roster_date_range(start_date, end_date):
    if not supabase: return
    try:
        res = supabase.table("worker").select("id").ilike("name", "___CONFIG___|%").execute()
        if res.data:
            supabase.table("worker").update({"name": f"___CONFIG___|{start_date}|{end_date}"}).eq("id", res.data[0]["id"]).execute()
        else:
            supabase.table("worker").insert({"name": f"___CONFIG___|{start_date}|{end_date}"}).execute()
    except Exception as e:
        print(f"Error saving config: {e}")

def get_dynamic_days():
    start_str, end_str = get_roster_date_range()
    today = date.today()
    if start_str and end_str:
        try:
            start_d = datetime.strptime(start_str, "%Y-%m-%d").date()
            end_d = datetime.strptime(end_str, "%Y-%m-%d").date()
        except ValueError:
            start_d, end_d = today, today + timedelta(days=6)
    else:
        start_d, end_d = today, today + timedelta(days=6)
    
    if end_d < start_d:
        end_d = start_d
        
    days = []
    curr = start_d
    while curr <= end_d:
        days.append(curr.strftime("%Y-%m-%d"))
        curr += timedelta(days=1)
    return days

def format_day_short(d_str):
    try:
        d = datetime.strptime(d_str, "%Y-%m-%d")
        return d.strftime("%d %a")
    except:
        return d_str


_DEFAULT_SHIFTS = [
    dict(code="AM",    label="AM Shift",    start_time="6:00 AM",  end_time="3:00 PM",  capacity=3, sort_order=0),
    dict(code="SWING", label="Swing Shift", start_time="11:00 AM", end_time="8:00 PM",  capacity=1, sort_order=1),
    dict(code="PM",    label="PM Shift",    start_time="3:00 PM",  end_time="12:00 AM", capacity=3, sort_order=2),
]

def init_db():
    if not supabase:
        return
    try:
        res = supabase.table("shift_config").select("id", count="exact").execute()
        if res.count == 0:
            supabase.table("shift_config").insert(_DEFAULT_SHIFTS).execute()
    except Exception as e:
        print(f"Error seeding shifts: {e}")

init_db()

# ------------------------------------------------------------- utilities --

def get_shift_configs():
    if not supabase: return []
    try:
        res = supabase.table("shift_config").select("*").order("sort_order").execute()
        return [SimpleNamespace(**d) for d in res.data]  # type: ignore[arg-type]
    except Exception as e:
        print(f"Error fetching shift configs: {e}")
        return []

def shift_meta_dict():
    return {s.code: s for s in get_shift_configs()}

def worker_leave_set(worker):
    """Return set of leave day names for a worker."""
    if not getattr(worker, 'leave_days', None):
        return set()
    return {d.strip() for d in worker.leave_days.split(",") if d.strip()}

def get_all_workers_with_assignments():
    if not supabase: return []
    try:
        res = supabase.table("worker").select("*, assignments:assignment(*)").not_.ilike("name", "___CONFIG___|%").order("name").execute()
        workers = []
        for d in res.data:
            w = SimpleNamespace(**d)  # type: ignore[arg-type]
            w.assignments = [SimpleNamespace(**a) for a in getattr(w, 'assignments', [])]
            workers.append(w)
        return workers
    except Exception as e:
        print(f"Error fetching workers: {e}")
        return []


def grouped_workers_for_day(day_str):
    """Return (groups, unassigned, off_workers, leave_workers, all_workers) for a given date."""
    configs    = get_shift_configs()
    codes      = [c.code for c in configs]
    all_workers = get_all_workers_with_assignments()

    groups        = {code: [] for code in codes}
    unassigned    = []
    off_workers   = []
    leave_workers = []

    for w in all_workers:
        assign = next((a for a in w.assignments if a.day == day_str), None)
        if assign:
            if assign.shift_code in groups:
                groups[assign.shift_code].append(w)
            elif assign.shift_code == "OFF":
                off_workers.append(w)
            elif assign.shift_code == "LEAVE":
                leave_workers.append(w)
            else:
                unassigned.append(w)
        else:
            # Check leave day by weekday name for backwards compatibility
            try:
                dt = datetime.strptime(day_str, "%Y-%m-%d")
                weekday_name = dt.strftime("%A")
            except:
                weekday_name = day_str
            if weekday_name in worker_leave_set(w):
                leave_workers.append(w)
            else:
                unassigned.append(w)

    return groups, unassigned, off_workers, leave_workers, all_workers

def name_exists(name):
    if not supabase: return False
    res = supabase.table("worker").select("id").ilike("name", name.strip()).execute()
    return len(res.data) > 0


# ----------------------------------------------------------------- views --

@app.route("/")
def index():
    configs = get_shift_configs()

    view_mode = request.args.get("view", "weekly")
    if view_mode not in ("daily", "weekly"):
        view_mode = "weekly"

    days_list = get_dynamic_days()
    days_formatted = {d: format_day_short(d) for d in days_list}
    
    current_day = request.args.get("day")
    if current_day not in days_list:
        current_day = days_list[0] if days_list else date.today().strftime("%Y-%m-%d")

    groups, unassigned, off_workers, leave_workers, all_workers = grouped_workers_for_day(current_day)

    # Build weekly grid data
    weekly_schedule = []
    for w in all_workers:
        assign_map = {a.day: (a.shift_code or "—") for a in w.assignments}
        leaves = worker_leave_set(w)
        row = {"worker": w, "days": {}}
        for d in days_list:
            if d in assign_map:
                row["days"][d] = assign_map[d]
            else:
                try:
                    weekday_name = datetime.strptime(d, "%Y-%m-%d").strftime("%A")
                except:
                    weekday_name = d
                if weekday_name in leaves:
                    row["days"][d] = "LEAVE"
                else:
                    row["days"][d] = "—"
        weekly_schedule.append(row)

    return render_template(
        "index.html",
        configs=configs,
        groups=groups,
        unassigned=unassigned,
        off_workers=off_workers,
        leave_workers=leave_workers,
        shift_meta=shift_meta_dict(),
        total=len(all_workers),
        today=date.today().strftime("%A, %d %B %Y"),
        today_iso=date.today().strftime("%Y-%m-%d"),
        days=days_list,
        days_formatted=days_formatted,
        current_day=current_day,
        view_mode=view_mode,
        weekly_schedule=weekly_schedule,
    )


@app.route("/allocate", methods=["POST"])
def allocate():
    if not supabase:
        flash("Supabase not configured.", "danger")
        return redirect(url_for("index"))

    workers = get_all_workers_with_assignments()
    if not workers:
        flash("Add some workers first, then allocate shifts.", "warning")
        return redirect(url_for("index"))

    configs = get_shift_configs()
    days_list = get_dynamic_days()

    # Wipe existing assignments for the CURRENT date range to preserve history
    if days_list:
        supabase.table("assignment").delete().in_("day", days_list).execute()

    workers_list = list(workers)
    random.shuffle(workers_list)  # initial fairness shuffle

    # ── Pre-compute leave sets once ──────────────────────────────────────
    leave_map = {w.id: worker_leave_set(w) for w in workers_list}

    new_assignments = []

    # ── Step 1: Insert LEAVE rows ────────────────────────────────────────
    for w in workers_list:
        for d in leave_map[w.id]:
            new_assignments.append({"worker_id": w.id, "day": d, "shift_code": "LEAVE"})

    # ── Step 2: Spread OFF days evenly across the week ───────────────────
    # Always pick the day that currently has the fewest offs so that no
    # single day ends up short-staffed (6/7 days worked per person).
    day_off_count = {d: 0 for d in days_list}
    off_day_map = {}   # worker_id -> off_day

    for w in workers_list:
        available_days = []
        for d in days_list:
            try:
                weekday_name = datetime.strptime(d, "%Y-%m-%d").strftime("%A")
            except:
                weekday_name = d
            if weekday_name not in leave_map[w.id]:
                available_days.append(d)
        
        if not available_days:
            continue  # fully on leave
        # Pick day with fewest offs; random tiebreak for variety
        best = min(available_days, key=lambda d: (day_off_count[d], random.random()))
        off_day_map[w.id] = best
        day_off_count[best] += 1
        new_assignments.append({"worker_id": w.id, "day": best, "shift_code": "OFF"})

    # ── Step 3: Day-by-day shift assignment with cross-week balancing ─────
    # shift_tally tracks how many of each shift each worker has had this
    # week so far.  When filling a slot, the worker with the fewest
    # assignments of that shift type wins → shifts rotate naturally.
    shift_tally = {
        w.id: {cfg.code: 0 for cfg in configs} for w in workers_list
    }

    for d in days_list:
        # Workers available to work today
        active = []
        for w in workers_list:
            try:
                weekday_name = datetime.strptime(d, "%Y-%m-%d").strftime("%A")
            except:
                weekday_name = d
            if weekday_name not in leave_map[w.id] and off_day_map.get(w.id) != d:
                active.append(w)

        # Build slot list and shuffle so no shift type always drafts first
        slots = []
        for cfg in configs:
            slots.extend([cfg.code] * cfg.capacity)
        random.shuffle(slots)

        pool = list(active)   # workers still unassigned today
        random.shuffle(pool)  # random start for fair tiebreaks

        for shift_code in slots:
            if not pool:
                break
            # Worker with fewest assignments of this shift gets it;
            # random tiebreak ensures variety run-to-run
            pool.sort(key=lambda w: (shift_tally[w.id][shift_code], random.random()))
            chosen = pool.pop(0)
            new_assignments.append({"worker_id": chosen.id, "day": d, "shift_code": shift_code})
            shift_tally[chosen.id][shift_code] += 1

        # Workers left in pool have no slot this day → idle/unassigned
        for w in pool:
            new_assignments.append({"worker_id": w.id, "day": d, "shift_code": None})

    if new_assignments:
        supabase.table("assignment").insert(new_assignments).execute()

    flash("Shifts allocated — 1 day off per worker, workload and shifts balanced.", "success")

    view = request.form.get("view", "weekly")
    cday = request.form.get("day", days_list[0] if days_list else date.today().strftime("%Y-%m-%d"))
    return redirect(url_for("index", view=view, day=cday))


@app.route("/reset", methods=["POST"])
def reset():
    if supabase:
        days_list = get_dynamic_days()
        if days_list:
            supabase.table("assignment").delete().in_("day", days_list).execute()
    flash("Shift assignments for the current date range cleared.", "info")
    view = request.form.get("view", "weekly")
    cday = request.form.get("day", days_list[0] if days_list else date.today().strftime("%Y-%m-%d"))
    return redirect(url_for("index", view=view, day=cday))


@app.route("/assign", methods=["POST"])
def assign_shift():
    """AJAX endpoint: update a single worker/day shift assignment."""
    if not supabase:
        return jsonify({"ok": False, "error": "Supabase not configured."}), 500

    data = request.get_json(force=True)
    worker_id  = data.get("worker_id")
    day        = data.get("day")
    shift_code = data.get("shift_code")  # may be None / "" for unassigned

    try:
        worker_id = int(worker_id)
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "Invalid worker_id."}), 400

    if not worker_id or not day:
        return jsonify({"ok": False, "error": "Invalid parameters."}), 400

    # Normalise empty string to None
    if shift_code == "" or shift_code == "—":
        shift_code = None

    try:
        # Check if an assignment row already exists
        res = supabase.table("assignment") \
            .select("id") \
            .eq("worker_id", worker_id) \
            .eq("day", day) \
            .execute()

        if res.data:
            row_id = res.data[0]["id"]
            supabase.table("assignment") \
                .update({"shift_code": shift_code}) \
                .eq("id", row_id) \
                .execute()
        else:
            supabase.table("assignment") \
                .insert({"worker_id": worker_id, "day": day, "shift_code": shift_code}) \
                .execute()

        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# -------------------------------------------------------- settings page --

@app.route("/settings")
def settings():
    configs = get_shift_configs()
    workers = get_all_workers_with_assignments()
    start_date, end_date = get_roster_date_range()
    return render_template("settings.html", configs=configs, workers=workers, 
                           start_date=start_date, end_date=end_date, days=["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"])


@app.route("/settings/save", methods=["POST"])
def settings_save():
    configs = get_shift_configs()
    errors  = []
    for cfg in configs:
        prefix     = f"shift_{cfg.code}_"
        label      = request.form.get(prefix + "label",      "").strip()
        start_time = request.form.get(prefix + "start_time", "").strip()
        end_time   = request.form.get(prefix + "end_time",   "").strip()
        cap_str    = request.form.get(prefix + "capacity",   "1").strip()

        if not label:
            errors.append(f"Label for {cfg.code} cannot be empty.")
            continue
        try:
            cap = int(cap_str)
            if cap < 1:
                raise ValueError
        except ValueError:
            errors.append(f"Capacity for {cfg.code} must be a positive integer.")
            continue
            
        if supabase:
            supabase.table("shift_config").update({
                "label": label,
                "start_time": start_time,
                "end_time": end_time,
                "capacity": cap
            }).eq("id", cfg.id).execute()

    # Roster Date Range Save
    roster_start = request.form.get("roster_start", "").strip()
    roster_end = request.form.get("roster_end", "").strip()
    if roster_start and roster_end:
        set_roster_date_range(roster_start, roster_end)

    if errors:
        for e in errors:
            flash(e, "warning")
    else:
        flash("Settings saved successfully.", "success")

    return redirect(url_for("settings"))


@app.route("/settings/leave", methods=["POST"])
def settings_leave():
    workers = get_all_workers_with_assignments()
    WEEKDAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    for w in workers:
        chosen = [d for d in WEEKDAYS if request.form.get(f"leave_{w.id}_{d}")]
        leave_str = ",".join(chosen)
        if supabase:
            supabase.table("worker").update({"leave_days": leave_str}).eq("id", w.id).execute()
            
    flash("Leave settings saved successfully.", "success")
    return redirect(url_for("settings"))


# ------------------------------------------------------- workers page --

@app.route("/workers")
def workers():
    all_workers = get_all_workers_with_assignments()
    # For the badge column use today as a reference day
    groups, _, _, _, _ = grouped_workers_for_day(date.today().strftime("%Y-%m-%d"))
    return render_template(
        "workers.html",
        all_workers=all_workers,
        shift_meta=shift_meta_dict(),
    )


@app.route("/workers/add", methods=["POST"])
def add_worker():
    name = request.form.get("name", "").strip()
    if not name:
        flash("Enter a name before adding a worker.", "warning")
    elif name_exists(name):
        flash(f'"{name}" is already on the list.', "warning")
    else:
        if supabase:
            try:
                supabase.table("worker").insert({"name": name}).execute()
                flash(f'"{name}" added.', "success")
            except Exception as e:
                flash(f"Database error adding worker: {e}", "danger")
        else:
            flash("Supabase is not configured.", "danger")
    return redirect(url_for("workers"))


@app.route("/workers/delete/<int:worker_id>", methods=["POST"])
def delete_worker(worker_id):
    if supabase:
        try:
            res = supabase.table("worker").select("name").eq("id", worker_id).execute()
            if res.data:
                name = res.data[0]["name"]
                supabase.table("worker").delete().eq("id", worker_id).execute()
                flash(f'"{name}" removed.', "info")
            else:
                flash("Worker not found.", "warning")
        except Exception as e:
            flash(f"Database error deleting worker: {e}", "danger")
    return redirect(url_for("workers"))



@app.route("/workers/upload", methods=["POST"])
def upload_workers():
    file = request.files.get("file")
    if not file or file.filename == "":
        flash("Choose an Excel file to upload.", "warning")
        return redirect(url_for("workers"))

    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        flash("Please upload a .xlsx file.", "warning")
        return redirect(url_for("workers"))

    try:
        wb    = load_workbook(file, read_only=True, data_only=True)
        sheet = wb.active
    except Exception:
        flash("Could not read that file. Make sure it's a valid Excel workbook.", "warning")
        return redirect(url_for("workers"))

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        flash("That sheet looks empty.", "warning")
        return redirect(url_for("workers"))

    header   = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    name_col = header.index("name") if "name" in header else 0
    data_rows = rows[1:] if "name" in header else rows

    added, skipped = 0, 0
    seen = set()
    
    new_workers = []
    
    for row in data_rows:
        if not row or name_col >= len(row):
            continue
        raw = row[name_col]
        if raw is None:
            continue
        name = str(raw).strip()
        if not name:
            continue
        key = name.lower()
        if key in seen or name_exists(name):
            skipped += 1
            continue
        seen.add(key)
        new_workers.append({"name": name})
        added += 1

    if new_workers and supabase:
        supabase.table("worker").insert(new_workers).execute()

    msg = f"Imported {added} worker(s)."
    if skipped:
        msg += f" Skipped {skipped} duplicate/blank row(s)."
    flash(msg, "success" if added else "warning")
    return redirect(url_for("workers"))


@app.route("/template")
def download_template():
    wb    = Workbook()
    sheet = wb.active
    sheet.title = "Workers"
    sheet["A1"] = "Name"
    sheet["A1"].font = Font(name="Arial", bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="2B2F4C")
    sheet["A2"] = "Jane Doe"
    sheet["A2"].font = Font(name="Arial", italic=True, color="8B90A8")
    sheet.column_dimensions["A"].width = 28
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="worker_import_template.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/export")
def export_schedule():
    all_workers = get_all_workers_with_assignments()

    wb    = Workbook()
    sheet = wb.active
    sheet.title = "Weekly Schedule"

    FILL = {
        "AM":    PatternFill("solid", fgColor="F2B872"),
        "SWING": PatternFill("solid", fgColor="E8734A"),
        "PM":    PatternFill("solid", fgColor="5C6BC0"),
        "OFF":   PatternFill("solid", fgColor="E2E4E9"),
        "LEAVE": PatternFill("solid", fgColor="A6ADB5"),
    }
    DARK_TEXT = {"PM", "LEAVE"}

    title_font  = Font(name="Arial", bold=True, size=14, color="1B2036")
    hdr_font    = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    hdr_fill    = PatternFill("solid", fgColor="2B2F4C")
    body_font   = Font(name="Arial", size=10)
    white_font  = Font(name="Arial", size=10, color="FFFFFF")
    center      = Alignment(horizontal="center", vertical="center")
    left        = Alignment(horizontal="left",   vertical="center")
    thin        = Side(style="thin", color="D9D9D9")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    sheet.merge_cells("A1:H1")
    sheet["A1"] = f"Weekly Shift Schedule — {date.today().strftime('%d %B %Y')}"
    sheet["A1"].font = title_font
    sheet["A1"].alignment = left
    sheet.row_dimensions[1].height = 28

    days_list = get_dynamic_days()
    headers = ["Worker Name"] + [format_day_short(d) for d in days_list]
    sheet.row_dimensions[3].height = 24
    for ci, h in enumerate(headers, start=1):
        c = sheet.cell(row=3, column=ci, value=h)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = center; c.border = border

    sheet.column_dimensions["A"].width = 26
    for col in ["B","C","D","E","F","G","H"]:
        sheet.column_dimensions[col].width = 14

    for ri, w in enumerate(all_workers, start=4):
        sheet.row_dimensions[ri].height = 20
        nc = sheet.cell(row=ri, column=1, value=w.name)
        nc.font = body_font; nc.alignment = left; nc.border = border

        assign_map = {a.day: (a.shift_code or "—") for a in w.assignments}
        leaves     = worker_leave_set(w)

        for ci, day_str in enumerate(days_list, start=2):
            if day_str in assign_map:
                val = assign_map[day_str]
            else:
                try:
                    weekday_name = datetime.strptime(day_str, "%Y-%m-%d").strftime("%A")
                except:
                    weekday_name = day_str
                if weekday_name in leaves:
                    val = "LEAVE"
                else:
                    val = "—"

            cell = sheet.cell(row=ri, column=ci, value=val)
            cell.alignment = center; cell.border = border
            if val in FILL:
                cell.fill = FILL[val]
                cell.font = white_font if val in DARK_TEXT else body_font
            else:
                cell.font = body_font

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"weekly_schedule_{date.today().isoformat()}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


if __name__ == "__main__":
    app.run(debug=True)
